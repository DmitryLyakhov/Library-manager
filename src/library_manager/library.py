#!/user/bin/env python3

import openpyxl
import pathlib
import heapq
from pathlib import Path
import json

"""Домашняя библиотека"""   

class Book:
    """Репрезентация книги в библиотеке"""
    json_path = Path(__file__).parent / 'genres.json'
    with open("book_chars.json", "r") as file:
        book_chars = json.load(file)
    GENRES = book_chars["GENRES"]
    SUBGENRES = book_chars["SUBGENRES"]
    FORMS = book_chars["FORMS"]
    def __init__(self, title, author, form="", genre="", subgenre="", rating=0, year=0, review=""):
        Book.CheckingRating(rating)
        Book.CheckingGenre(genre)
        Book.CheckingSubgenre(subgenre)
        Book.CheckingForm(form)
        self.title = title
        self.author = author
        self.form = form
        self.genre = genre
        self.subgenre = subgenre
        self.rating = rating
        self.year = year
        self.review = review
    def __repr__(self):
        return (f"Title - {self.title}, author - {self.author}, form - {self.form}, " 
        f"genre - {self.genre}, rating - {self.rating}, year - {self.year}")
    def __str__(self):
        return (f"{self.title}, за авторством {self.author}. {self.genre if self.genre not in ("", None) else 'жанр(ы) не указан(ы)'}, " 
        f"{self.subgenre if self.subgenre not in ("", None) else 'поджанр(ы) не указан(ы)'}; " f"{self.form if self.form not in ("", None) else 'форма не указана'}. Оценка - {self.rating}. "
        f"Год - {self.year if self.year != 0 else 'не указан'}\nотзыв - {self.review if self.review not in ("", None) else 'не указан'}")
    def title_update(self, title):
        self.title = title
    def author_update(self, author):
        self.author = author
    def form_update(self, form):
        Book.CheckingForm(form)
        self.form = form
    def genre_update(self, genre):
        Book.CheckingGenre(genre)
        self.genre = genre
    def rating_update(self, rating):
        Book.CheckingRating(rating)
        self.rating = rating
    def year_update(self, year):
        self.year = year
    def review_update(self, review):
        self.review = review
    def subgenre_update(self, subgenre):
        Book.CheckingSubgenre(subgenre)
        self.subgenre = subgenre
    def CheckingRating(rating):
        if not isinstance(rating, (int, float)) or (rating>10 or rating<0):
            print("Оценка должна быть в диапазоне от 0 до 10 включительно")
    def CheckingGenre(genre):
        if genre is None:
            genre = ""
        if genre != "":
            genre_splitted = [el.strip(" ") for el in genre.split(",")]
            if len(genre_splitted) >= 2:
                for elem in genre_splitted:
                    if elem not in Book.GENRES:
                        print("Неизвесный(е) жанр(ы)")
            else:
                if genre_splitted[0] not in Book.GENRES and genre != "":
                    print("Неизвесный(е) жарнр(ы)")
    def CheckingSubgenre(subgenre):
        if subgenre is None:
            subgenre = ""
        if subgenre != "":
            subgenre_splitted = [el.strip(" ") for el in subgenre.split(",")]
            if len(subgenre_splitted) >= 2:
                for elem in subgenre_splitted:
                    if elem not in Book.GENRES:
                        print("Неизвесный(е) поджанр(ы)")
            else:
                if subgenre_splitted[0] not in Book.SUBGENRES and subgenre != "":
                    print("Неизвесный(е) поджарнр(ы)")
    def CheckingForm(form):
        if form is None:
            form = ""
        if form not in Book.FORMS and form != "":
            print("Неизвестная форма")
    @classmethod
    def update_chars(cls):
        chars_path = Path(__file__).parent / 'book_chars.json'
        with open(chars_path, "r") as file:
            book_chars = json.load(file)
        cls.GENRES = book_chars["GENRES"]
        cls.SUBGENRES = book_chars["SUBGENRES"]
        cls.FORMS = book_chars["FORMS"]
        
  
class Library: 
    lib_path = pathlib.Path.home() / "Desktop/home_library"
    book_dummy = Book("a", "b")
    num_of_book_params = len(list(vars(book_dummy).keys()))
    def __init__(self, name):
        self.name = name
        pathlib.Path(Library.lib_path).mkdir(exist_ok=True)
        if (Library.lib_path / f"{name}.xlsx").is_file():
            pass
        else:
            library = openpyxl.Workbook()
            ws = library.active
            for i in range (1, Library.num_of_book_params+1):
                ws.cell(1, i).value = list(vars(Library.book_dummy).keys())[i-1]
            library.save(Library.lib_path / f"{self.name}.xlsx")
    def book_add(self, book):
        wb, ws = self.openlibrary()
        not_dupl = True
        for i in range(2, ws.max_row+1):
            if book.title == ws.cell(i, 1).value:
                book_title = book.title
                book_author = ws.cell(i, 2).value
                book_elements = [ws.cell(i, k).value for k in range(3, Library.num_of_book_params+1)]
                book_args = [ws.cell(1, k).value for k in range(3, Library.num_of_book_params+1)]
                other = dict(zip(book_args, book_elements))
                BookDupli = Library.create_book_obj(book_title, book_author, **other)
                not_dupl = False
                print(f"Книга с таким названием уже есть в библиотеке: {str(BookDupli)}")
        book_row = ws.max_row+1
        if not_dupl:
            for i in range(1, Library.num_of_book_params+1):
                ws.cell(book_row, i).value = list(vars(book).values())[i-1]
        wb.save(Library.lib_path / f"{self.name}.xlsx")
    def filter_books(self, exclusive=False, **kwargs):
        wb, ws = self.openlibrary()
        elements_of_search = list(kwargs.keys())
        values_to_search = list(kwargs.values())
        rows_needed = []
        books = []
        for i in range(1, Library.num_of_book_params+1):
            if ws.cell(1, i).value in elements_of_search:
                rows_needed.append(i)
        for i in range (2, ws.max_row+1):
            summa = 0
            counter = 0
            for elem in rows_needed:
                try: 
                    val_sep = [int(el.strip()) for el in values_to_search[counter].split(",")]
                except ValueError:
                    val_sep = [el.strip() for el in values_to_search[counter].split(",")]
                cell_value = ws.cell(i, elem).value
                if cell_value is None:
                    cell_values = []
                else:
                    try:
                        cell_values = [el.strip(" ") for el in ws.cell(i, elem).value.split(",")]
                    except AttributeError:
                        cell_values = [ws.cell(i, elem).value]
                if exclusive:
                    if sorted(cell_values) == sorted(val_sep):
                        summa += 1
                else:
                    if any(item in cell_values for item in val_sep):
                        summa += 1
                counter += 1
            if summa == len(elements_of_search):
                book_title = ws.cell(i, 1).value
                book_author = ws.cell(i, 2).value
                book_elements = [ws.cell(i, k).value for k in range(3, Library.num_of_book_params+1)]
                book_args = [ws.cell(1, k).value for k in range(3, Library.num_of_book_params+1)]
                other = dict(zip(book_args, book_elements))
                BookFetched = Library.create_book_obj(book_title, book_author, **other)
                books.append(BookFetched)
        wb.save(Library.lib_path / f"{self.name}.xlsx")
        return books
    def delete_book(self, **kwargs): #словарь где ключи - title, значения - author
        wb, ws = self.openlibrary()
        titles = list(kwargs.keys())
        authors = list(kwargs.values())
        indexes = []
        for cur_title in range(len(titles)):
            for i in range(1, ws.max_row+1):
                if ws.cell(i, 1).value == titles[cur_title] and ws.cell(i, 2).value == authors[cur_title]:
                    indexes.append(i)
        for row in indexes[::-1]:
            ws.delete_rows(row)
        wb.save(Library.lib_path / f"{self.name}.xlsx")
    def replace_book(self, book):
        wb, ws = Library.openlibrary(self)
        row = []
        for i in range(1, ws.max_row+1):
            if ws.cell(i,1).value == book.title and ws.cell(i,2).value == book.author:
                row.append(i)
        if len(row) > 1:
            print(f"Найдено больше одной книги с такими названием и автором, проверьте ряды {row} библиотеки")
        elif len(row) == 0:
            print("Такая книга не найдена в библиотеке")
        else:
            for i in range(1, Library.num_of_book_params+1):
                ws.cell(row[0], i).value = list(vars(book).values())[i-1]
        wb.save(Library.lib_path / f"{self.name}.xlsx")
    def finding_top(self, top=10, exclusive=False, **kwargs):
        books = self.filter_books(exclusive = exclusive, **kwargs)
        top_books = heapq.nlargest(top, books, key = lambda book: book.rating)
        for top in top_books:
            print(top)
    def openlibrary(self):
        wb = openpyxl.load_workbook(Library.lib_path / f"{self.name}.xlsx")
        ws = wb.active
        return wb, ws
    def create_book_obj(title, author, **kwargs):
        return Book(title, author, **kwargs)  

if __name__ == "__main__":
    book = Book("a", "b", form="стих")